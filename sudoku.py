from mypulp import *
from openpyxl import *
from copy import copy

book = load_workbook('sudoku_q.xlsx')
sheet = book['q1']

#データの作成
N = [i for i in range(1,10)]
C = [(3*i+2, 3*j+2) for i in range(3) for j in range(3)]
D = [(i-1, j-1) for i in range(3) for j in range(3)]

n = {}
for i in N:
    for j in N:
        n[i, j] = sheet.cell(row=i+1, column=j+3).value
model = Model()

#決定変数
x = {}
for i in N:
    for j in N:
        for k in N:
            x[i, j, k] = model.addVar(vtype='B')

#制約条件1
for i in N:
    for j in N:
        if type(n[i, j]) is int and 1 <= n[i, j] <= 9:
            model.addConstr(x[i,j,n[i, j]]==1)
        else:
            model.addConstr(quicksum(x[i,j,k] for k in N)==1)

#制約条件2
for i in N:
    for k in N:
        model.addConstr(quicksum(x[i,j,k] for j in N)==1)

#制約条件3
for j in N:
    for k in N:
        model.addConstr(quicksum(x[i,j,k] for i in N)==1)

#制約条件4
for (ci, cj) in C:
    for k in N:
        model.addConstr(quicksum(x[ci+di, cj+dj, k] for (di, dj) in D)==1)
        
#目的関数
model.setObjective(x[1,1,1], GRB.MAXIMIZE)

#実行
model.optimize()
if model.Status == GRB.Status.OPTIMAL:
    print('解が見つかりました。')
    #出力
    for i in N:
        for j in N:
            for k in N:
                if x[i, j, k].X > 0.01:
                    font = copy(sheet.cell(row=i+1, column=j+13).font)
                    if type(n[i, j]) is int:
                        font.color = styles.colors.Color(rgb='ffff0000')
                    else:
                        font.color = styles.colors.Color(rgb='ff000000')
                    sheet.cell(row=i+1, column=j+13).font = font
                    sheet.cell(row=i+1, column=j+13).value = k                    
    book.save('sudoku_ans.xlsx')
else:
    print('解が見つかりませんでした。')